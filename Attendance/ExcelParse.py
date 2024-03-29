from openpyxl import load_workbook
from Myconfig import *
import Data as globaldata
from persondata import *
import ConfigParse as configdata
import ctypes
def ParseAttendanceSheet(path,Debug=0):
    globaldata.AllDataList.clear()
    if Debug == 0:
        path = str(path)[3:-2].replace("\\\\","\\").replace("\\\\","\\")
    print(f"{bcolors.OKCYAN}{type(path)}考勤表路径为:{path}{bcolors.ENDC}")
    if "xlsx" not in path:
        ctypes.windll.user32.MessageBoxW(0, "请拖入Excel表格", "错误", 0)
        return
    temp = load_workbook(path).active
    if temp is None:
        ctypes.windll.user32.MessageBoxW(0, "表格数据有问题 请检查", "错误", 0)
        return
    for index in range(globaldata.BeginNumber,temp.max_row+1):
        NowIndex = str(index)
        name = temp[globaldata.NameKey + NowIndex]
        join = temp[globaldata.DateKey + NowIndex]

        onwork = temp[globaldata.OnWorkKey + NowIndex]
        offwork = temp[globaldata.OffWorkKey + NowIndex]

        person = PersonData(name.value, str(join.value), onwork.value, offwork.value)
        print(person)
        globaldata.AllDataList.append(person)

        if configdata.SkipDate(person.JoinDateTime):
            globaldata.SkipedList.append(person)
            continue
        # if  type(person.OnWorkTime) is str:
        #     globaldata.NoOnWorkTimeList.append(person)
        # elif IsNormalDayBeLate(person):
        #     globaldata.LateList.append(person)
        # if type(person.OffWorkTime) is str:
        #     globaldata.NoOffWorkTimeList.append(person)
        # elif(IsNormalDayLeaveEarly(person)):
        #     globaldata.LeaveEarlyList.append(person)

import tkinter
import sys
import tkinter.ttk
from tkcalendar import Calendar, DateEntry
import Data as globaldata
if sys.platform == 'WIN32':
    from Myconfig import *
from openpyxl import load_workbook
from tkinter import *
from persondata import *
import warnings
attendanceSheetPath=""
import windnd
import ctypes
# 跳过或者添加特定日期
def SkipDate(JoinDateTime:datetime):
    for date in globaldata.ContainInfoDate:
       if date == JoinDateTime:
           return False
    for date in globaldata.SkipInfoDate:
        if date == globaldata.SkipInfoDate:
            return True

    if JoinDateTime.weekday() == 6 or JoinDateTime.weekday() == 7:
        return True
    return False


def ParseAttendanceSheet(path):
    path = str(path)[3:-2].replace("\\\\","\\").replace("\\\\","\\")
    print(f"{bcolors.OKCYAN}{type(path)}考勤表路径为:{path}{bcolors.ENDC}")
    if "xlsx" not in path:
        ctypes.windll.user32.MessageBoxW(0, "请拖入Excel表格", "错误", 0)
    temp = load_workbook(path).active
    if temp is None:
        ctypes.windll.user32.MessageBoxW(0, "表格数据有问题 请检查", "错误", 0)
    for index in range(globaldata.BeginNumber,temp.max_row):
        NowIndex = str(index)
        name = temp[globaldata.NameKey + NowIndex]
        join = temp[globaldata.JoinKey + NowIndex]
        onwork = temp[globaldata.OnWorkKey + NowIndex]
        offwork = temp[globaldata.OffWorkKey + NowIndex]
        person=PersonData(name.value,join.value,onwork.value,offwork.value)
        person.print()
        globaldata.AllDataList.append(person)

        if SkipDate(person.JoinDateTime):
            globaldata.SkipedList.append(person)
            continue

        if  type(person.OnWorkTime) is str:
            globaldata.NoOnWorkTimeList.append(person)
        elif IsNormalDayBeLate(person):
            globaldata.LateList.append(person)
        if type(person.OffWorkTime) is str:
            globaldata.NoOffWorkTimeList.append(person)
        elif(IsNormalDayLeaveEarly(person)):
            globaldata.LeaveEarlyList.append(person)

def printname(fiel):
    print(fiel)


if __name__ == '__main__':
    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
    tk = Tk()
    right_frame = Frame(tk,bg="Grey",height=50,width=50)
    right_frame.grid(row=0, column=1,padx=20,pady=20)
    cal = Calendar(right_frame, font="Arial 14", selectmode='day', locale='zh_CN', disabledforeground='red',
                   cursor="hand1", year=2023, month=6, day=5)
    cal.pack(fill="both", expand=True)
    left_frame = Frame(tk,bg="Gray",height=50,width=50)
    left_frame.grid(row=0, column=0,padx=20,pady=20)
    if sys.platform == 'WIN32':
        windnd.hook_dropfiles(left_frame,func=ParseAttendanceSheet)
        windnd.hook_dropfiles(right_frame,func=printname)
    tk.mainloop()

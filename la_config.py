# -*- coding: utf-8 -*-

'通过config配置读取多语言文件 并输出结果到result.txt'

__author__ = 'tuFL'

import os
import sys
import re
from openpyxl import load_workbook
import json
from Myconfig import *

sheetmap = {}
configfile = []
configstr = []
UsingCount = 0
fileFD = []
resultfilename = "result.txt"


def LoadJson():
    xlsx = 'xlsx'
    str = 'str'
    f = open('LanguageConfig.json')
    config = json.load(f)
    f.close()
    print(f"{bcolors.HEADER} 需要读取的多语言表有:{config[xlsx]}{bcolors.ENDC}")
    print(f"{bcolors.OKBLUE} 需要匹配的数据:{config[str]}{bcolors.ENDC}")
    print(f"{bcolors.BOLD} 查询字数最多的:{config['count']}{bcolors.ENDC}")
    for i in config[xlsx]:
        configfile.append(i)
    for i in config[str]:
        configstr.append(i)
    global UsingCount
    UsingCount = config['count']


def LoadXlsl():
    Language_Path = "D:\\simulator\\simulator_landlord_GB\\ExDataHome\\Excel\\Excel";
    Lafile = []
    sheetlist = []
    os.chdir(Language_Path)
    for root, dirs, file in os.walk('.', topdown=False, followlinks=False):
        Lafile.extend([fi for fi in file if "Language" in fi and "csv" not in fi and '~' not in fi and any(
            name in fi for name in configfile)])

    for name in Lafile:
        temp = load_workbook(name).active
        sheetlist.append(temp)
        usedvale = temp['B6:' + 'B' + str(temp.max_row)]
        Values = [v.value for vale in usedvale for v in vale]
        usedvale = temp['A6:' + 'A' + str(temp.max_row)]
        Keys = [v.value for vale in usedvale for v in vale]

        sheetmap[temp.title] = dict(zip(Keys, Values))


def _Searchs(name, _str):
    dirc = sheetmap[name]
    templist = []
    for temp in dirc.items():
        if re.search(str(_str), str(temp[0])):
            templist.append(temp)

    templist = sorted(templist, key=lambda x: len(x[1]), reverse=True)
    fileFD.write(f"{name} [{_str}] \n")
    for tem in templist[:UsingCount]:
        fileFD.write(f"{tem[0]} {tem[1]} {len(tem[1])} \n")


def Searchs():
    for filename in configfile:
        for strs in configstr:
            _Searchs(filename, strs)


if __name__ == '__main__':
    nowpath = sys.path[0]
    os.chdir(nowpath)
    print(f"{bcolors.OKBLUE}Work path: {nowpath}{bcolors.ENDC}")
    if os.path.exists(resultfilename):
        os.remove(resultfilename)
    fileFD = open(resultfilename, "a+", encoding='utf-8')
    LoadJson()
    LoadXlsl()
    Searchs()
    fileFD.close()
    self_commands = f"EmEditor.exe {nowpath}\\{resultfilename}"
    os.system(self_commands)

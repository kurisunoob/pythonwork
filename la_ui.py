import os
import re
import sys
from time import strftime, localtime

import openpyxl.worksheet.worksheet
from openpyxl.workbook.child import _WorkbookChild
from Myconfig import *
from openpyxl import load_workbook
from tkinter import *
from tkinter import scrolledtext
from pygoogletranslation import Translator

sheetmap = {}
configfile = []
configstr = []
UsingCount = 0
fileFD = []
resultfilename = "result_ui"
ResultStr = ""
Editor = "EmEditor.exe"

FontList= {"Language_Arabic": "ARIALNB", "Language_Russian": "ARIALNB", "Language_TC": "Chinese_TC",
           "Language_DE": "English", "Language_English": "English", "Language_ES": "Vietnamese",
           "Language_French": "Vietnamese", "Language_Italian": "Vietnamese", "Language_PTBR": "Vietnamese",
           "Language_Vietnamese": "Vietnamese", "Language_Japanese": "Japanese", "Language_Korean": "Korean",
           "Language_Thai": "Thai","Language_TR": "Vietnamese"}


def SetPath():
    Language_Path = r"D:\simulator\simulator_landlord_GB\ExDataHome\Excel\Excel"
    os.chdir(Language_Path)


def LoadXlsl():
    Lafile = []
    sheetlist = []
    for root, dirs, file in os.walk('.', topdown=False, followlinks=False):
        for fi in file:
            if "Language" in fi and "csv" not in fi and '~' not in fi:
                for name in configfile:
                    if name in fi:
                        Lafile.append(fi)

    for name in Lafile:
        temp = load_workbook(name).active
        sheetlist.append(temp)
        tempdic = {}
        usedvale = temp['B6:' + 'B' + str(temp.max_row)]
        Values = []
        for vale in usedvale:
            for v in vale:
                Values.append(v.value)
        usedvale = temp['A6:' + 'A' + str(temp.max_row)]
        Keys = []
        for vale in usedvale:
            for v in vale:
                Keys.append(v.value)

        for index in range(len(Keys)):
            _key = Keys[index]
            _value = Values[index]
            tempdic[_key] = _value
        sheetmap[temp.title] = tempdic


def GetConfig():
    configfile.clear()
    configstr.clear()
    for strs in strtext.get().split(','):
        configstr.append(strs)
    global UsingCount
    UsingCount = str(strcount.get())
    for strs in CheckVarDic.items():
        if strs[1].get() == 1:
            configfile.append(strs[0].split(".")[0])
    LoadXlsl()


def Searchs():
    print(f"{bcolors.OKBLUE}Work path: {nowpath}{bcolors.ENDC}")
    if os.path.exists(resultfilename):
        os.remove(resultfilename)
    fileFD = open(resultfilename, "a+", encoding='utf-8')
    for filename in configfile:
        for strs in configstr:
            _Searchs(filename, strs, fileFD)
    fileFD.close()

def sorter(item):
    if item[1] is None :
        print(item)
        return False;
    return len(item[1])
def _Searchs(name, _str, _fileFD):
    dirc = sheetmap[name]
    templist = []
    templist.clear()
    for temp in dirc.items():
        if re.search(str(_str), str(temp[0])):
            templist.append(temp)

    templist = sorted(templist, key=sorter, reverse=True)

    global ResultStr
    _font=1;
    if name in FontList.keys():
        _font = FontList[name]
    _fileFD.write(f"{name} [{_str}]\n[{_font}] \n")
    ResultStr += (f"{name} [{_str}]\n[{_font}] \n")
    for tem in templist[:int(UsingCount)]:
        _fileFD.write(f"{tem[0]} {tem[1]} {len(tem[1])} \n")
        ResultStr += (f"{tem[0]}  {len(tem[1])}\n{tem[1]} \n")
def _Searchs_trans(name, _str, _fileFD):
    dirc = sheetmap[name]
    templist = []
    templist.clear()
    for temp in dirc.items():
        if re.search(str(_str), str(temp[0])):
            templist.append(temp)

    templist = sorted(templist, key=sorter, reverse=True)

    global ResultStr
    _font=1;
    if name in FontList.keys():
        _font = FontList[name]
    _fileFD.write(f"{name} [{_str}]\n[{_font}] \n")
    ResultStr += (f"{name} [{_str}]\n[{_font}] \n")
    for tem in templist[:int(UsingCount)]:
        trans = Translator()
        tp = trans.translate(tem[1], dest='zh-CN')
        trantext = tp.text
        _fileFD.write(f"{tem[0]} {tem[1]} {trantext} {len(tem[1])} \n")
        ResultStr += (f"{tem[0]}  {len(tem[1])}\n{tem[1]} \n")


def ButtonClick(Mod):
    global ResultStr
    ResultStr = ""
    GetConfig()
    Searchs()
    if (Mod == "App"):
        self_commands = f"{Editor} {resultfilename}"
        os.system(self_commands)
    elif (Mod == "Show"):
        # Label(root, text=ResultStr).grid(row=1, column=1)
        newwindow = Tk()
        window = scrolledtext.ScrolledText(newwindow, wrap=WORD, width=100, height=60, font=("Time New Roman", 15))
        window.insert(INSERT, ResultStr)
        window.grid(row=0, column=0)
        newwindow.mainloop()

def ButtonClick_trans():
    global ResultStr
    ResultStr = ""
    GetConfig()
    print(f"{bcolors.OKBLUE}Work path: {nowpath}{bcolors.ENDC}")
    if os.path.exists(resultfilename):
        os.remove(resultfilename)
    fileFD = open(resultfilename, "a+", encoding='utf-8')
    for filename in configfile:
        for strs in configstr:
            _Searchs_trans(filename, strs, fileFD)
    fileFD.close()
    self_commands = f"{Editor} {resultfilename}"
    os.system(self_commands)


def vaildcallback(event):
    print(event.char)
    if event.char.isdigit() or event.char == '\x08':
        return True
    return 'break'


def ClearTXT():
    Language_Path = "D:\\simulator\\PureObject\\simulator_landlord_GB\\ExDataHome\\Excel\\Excel"
    os.chdir(sys.path[0])
    for root, dirs, file in os.walk('.', topdown=False, followlinks=False):
        for fi in file:
            if "txt" in fi and "result" in fi:
                print(f"{bcolors.OKGREEN}Remove File{fi}{bcolors.ENDC}")
                os.remove(fi)
    os.chdir(Language_Path)
def ChooseLanguageAll():
    for Key in CheckVarDic:
            if CheckVarDic[Key].get() == 1:
                CheckVarDic[Key].set(0)
            else:
                CheckVarDic[Key].set(1)
def ChooseLanguageWB():
    for Key in CheckVarDic:
        if "Arabic" in Key \
                or "DE" in Key \
                or "Italian" in Key \
                or "PTBR" in Key \
                or "Russian" in Key \
                or "Thai" in Key \
                or "ES" in Key \
                 or "TR" in Key \
                 or "JP" in Key \
                 or "TC" in Key \
                or "Vietnamese" in Key:
            if CheckVarDic[Key].get() == 1:
                CheckVarDic[Key].set(0)
            else:
                CheckVarDic[Key].set(1)

def FormatCheck():
    global ResultStr
    ResultStr = ""
    GetConfig()
    Searchs_format()
    self_commands = f"{Editor} {resultfilename}"
    os.system(self_commands)


def Searchs_format():
    print(f"{bcolors.OKBLUE}Work path: {nowpath}{bcolors.ENDC}")
    if os.path.exists(resultfilename):
        os.remove(resultfilename)
    fileFD = open(resultfilename, "a+", encoding='utf-8')
    for filename in configfile:
        for strs in configstr:
            _Searchs_format(filename, strs, fileFD)
    fileFD.close()

def _Searchs_format(name, _str, _fileFD):
    dirc = sheetmap[name]
    templist = []
    templist.clear()
    for temp in dirc.items():
        if re.search(str(_str), str(temp[1])):
            templist.append(temp)

    templist = sorted(templist, key=sorter, reverse=True)

    global ResultStr
    _font=1;
    if name in FontList.keys():
        _font = FontList[name]
    _fileFD.write(f"{name} [{_str}]\n[{_font}] \n")
    ResultStr += (f"{name} [{_str}]\n[{_font}] \n")
    for tem in templist[:int(UsingCount)]:
        _fileFD.write(f"{tem[0]} {tem[1]} {len(tem[1])} \n")
        ResultStr += (f"{tem[0]}  {len(tem[1])}\n{tem[1]} \n")

def CheckTextQuantity():
    configfilelist=[]
    Lafile = []
    quantityset=set()

    for strs in CheckVarDic.items():
        if strs[1].get() == 1:
            configfilelist.append(strs[0].split(".")[0])
    for root, dirs, file in os.walk('.', topdown=False, followlinks=False):
        for fi in file:
            if "Language" in fi and "csv" not in fi and '~' not in fi:
                for name in configfilelist:
                    if name in fi:
                        Lafile.append(fi)

    for name in Lafile:
        temp = load_workbook(name).active
        quantityset.add(temp.max_row)
        if len(quantityset) > 1:
            print(f"{bcolors.FAIL}∑(￣ﾛ￣||)这个有问题{bcolors.ENDC}===>{bcolors.HEADER}{name}{bcolors.ENDC}的文本量为:{bcolors.OKBLUE}{temp.max_row}{bcolors.ENDC}")
            quantityset.remove(temp.max_row)
        else:
            print(f"{bcolors.HEADER}{name}{bcolors.ENDC}的文本量为:{bcolors.OKBLUE}{temp.max_row}{bcolors.ENDC}")
    print(f'{bcolors.OKGREEN}检查完成{bcolors.ENDC}')


if __name__ == '__main__':
    nowpath = sys.path[0]
    resultfilename = f"{nowpath}\\{resultfilename}{strftime('%Y_%m_%d_%H_%M_%S', localtime())}.txt"
    SetPath()
    CheckVarDic = {}
    root = Tk()
    root.title("te")
    root.config(bg="skyblue")
    # ======================================
    left_frame = LabelFrame(root, text="language xlsx")
    left_frame.grid(row=0, column=0, padx=20, pady=10)
    for roots, dirs, file in os.walk('.', topdown=False, followlinks=False):
        for fi in file:
            if "Language" in fi and "csv" not in fi and '~' not in fi:
                if 'CHS' in fi:
                    continue
                CheckVar = IntVar()
                ck = Checkbutton(master=left_frame, text=fi, variable=CheckVar, onvalue=1, offvalue=0, height=1,
                                 width=30, anchor="w")
                ck.pack()
                CheckVarDic[fi] = CheckVar
    # ======================================
    right_frame = Frame(root)
    right_frame.grid(row=0, column=1)
    Label(right_frame, text="匹配文本:").grid(row=0)
    Label(right_frame, text="展示数量:").grid(row=1)
    strtext = Entry(right_frame, width=30)
    strtext.insert(0, "TreeHole_Q_*,TreeHole_Answer_*")
    strtext.grid(row=0, column=1)
    strcount = Entry(right_frame, width=5)
    strcount.bind('<KeyPress>', vaildcallback)
    strcount.insert(0, "5")
    strcount.grid(row=1, column=1)

    # ====================================
    down_frame = Frame(root)
    down_frame.grid(row=1, column=0)
    Button(down_frame, text="查询_外部打开(查询Key)", command=lambda: ButtonClick("App")).pack()
    Button(down_frame, text="检查格式(查询文本)", command=lambda: FormatCheck()).pack()
    # Button(down_frame, text="查询_ 查看", command=lambda: ButtonClick("Show")).pack()
    Button(down_frame, text="一键选中外包翻译", command=lambda: ChooseLanguageWB()).pack()
    Button(down_frame, text="一键选中所有文本", command=lambda: ChooseLanguageAll()).pack()
    Button(down_frame, text="查看翻译(过翻译API速度慢)", command=lambda: ButtonClick_trans()).pack()
    Button(down_frame, text="检查文本数量", command=lambda: CheckTextQuantity()).pack()
    Button(down_frame, text="清理文本", command=lambda: ClearTXT()).pack()
    root.mainloop()
